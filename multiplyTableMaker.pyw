# user/bin/env python3
# -*- coding: utf-8 -*-
# multiplyTableMaker.pyw - Créer une table de multiplication. 

import openpyxl, os, sys
from openpyxl.cell import get_column_letter

os.chdir(r'/Applications/ATBS')

wb = openpyxl.Workbook()
sheet = wb.active

def columnAdata():
    # Create data in Column A.
    for i in range(1, multiplyNum + 1):
        sheet['A' + str(i + 1)] = i

def rowAdata():
    # Create data along Row 1.
    for j in range(1, multiplyNum + 1):
        sheet[get_column_letter(j + 1) + str(1)] = j

multiplyNum = int(sys.argv[1])
columnAdata()
rowAdata()

#Create the multiplication table from the input number from the command line.
for i in range(1, (multiplyNum + 1)):
    for j in range(1, (multiplyNum + 1)):
        for rowOfCellObjects in sheet['B2':get_column_letter(multiplyNum + 1) + str(multiplyNum + 1)]:
            for cellObj in rowOfCellObjects:
                # Calculate Multiplication
                sheet[get_column_letter(j + 1) + str(i + 1)] = i * j

wb.save('mutiplyOutput.xlsx')